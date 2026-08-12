from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def EPexportarVentasPDF(EPventas, EPruta):
    EPdocumento = SimpleDocTemplate(EPruta, pagesize=letter)
    EPestilos = getSampleStyleSheet()
    EPelementos = []
    EPelementos.append(Paragraph("Reporte de ventas - Panaderia", EPestilos["Title"]))
    EPelementos.append(Spacer(1, 0.5 * cm))
    EPencabezados = ["Fecha", "Producto", "Vendedor", "Cantidad", "Total"]
    EPfilas = [EPencabezados]
    EPtotalGeneral = 0
    for EPventa in EPventas:
        EPfecha = EPventa["fecha_hora"]
        EPfechaTexto = EPfecha.strftime("%d/%m/%Y %H:%M") if hasattr(EPfecha, "strftime") else str(EPfecha)
        EPfilas.append([
            EPfechaTexto, EPventa["nombre_producto"], EPventa["nombre_vendedor"],
            str(EPventa["cantidad"]), f"${float(EPventa['total']):.2f}"])
        EPtotalGeneral += float(EPventa["total"])
    EPfilas.append(["", "", "", "Total:", f"${EPtotalGeneral:.2f}"])
    EPtabla = Table(EPfilas, colWidths=[3.2 * cm, 4 * cm, 3.5 * cm, 2.3 * cm, 2.5 * cm])
    EPtabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8B5E3C")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -2), 0.5, colors.grey),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("LINEABOVE", (0, -1), (-1, -1), 1, colors.black),
        ("ALIGN", (3, 0), (4, -1), "RIGHT"),
    ]))
    EPelementos.append(EPtabla)
    EPdocumento.build(EPelementos)

def EPexportarVentasExcel(EPventas, EPruta):
    EPlibro = openpyxl.Workbook()
    EPhoja = EPlibro.active
    EPhoja.title = "Ventas"
    EPencabezados = ["Fecha", "Producto", "Vendedor", "Cantidad", "Total"]
    EPhoja.append(EPencabezados)
    for EPcelda in EPhoja[1]:
        EPcelda.font = Font(bold=True, color="FFFFFF")
        EPcelda.fill = PatternFill("solid", fgColor="8B5E3C")
        EPcelda.alignment = Alignment(horizontal="center")
    EPtotalGeneral = 0
    for EPventa in EPventas:
        EPfecha = EPventa["fecha_hora"]
        EPfechaTexto = EPfecha.strftime("%d/%m/%Y %H:%M") if hasattr(EPfecha, "strftime") else str(EPfecha)
        EPhoja.append([
            EPfechaTexto, EPventa["nombre_producto"], EPventa["nombre_vendedor"],
            EPventa["cantidad"], float(EPventa["total"])])
        EPtotalGeneral += float(EPventa["total"])
    EPhoja.append(["", "", "", "Total:", EPtotalGeneral])
    EPfilaTotal = EPhoja.max_row
    EPhoja.cell(row=EPfilaTotal, column=4).font = Font(bold=True)
    EPhoja.cell(row=EPfilaTotal, column=5).font = Font(bold=True)
    EPanchos = [18, 24, 18, 10, 12]
    for EPindice, EPancho in enumerate(EPanchos, start=1):
        EPhoja.column_dimensions[chr(64 + EPindice)].width = EPancho
    EPlibro.save(EPruta)